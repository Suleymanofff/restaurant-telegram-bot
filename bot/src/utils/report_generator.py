import os
import logging
from datetime import datetime, date
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile

logger = logging.getLogger(__name__)

class ReportGenerator:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    async def generate_daily_reservations_report(self, reservations: List[Dict[str, Any]]) -> str:
        """
        Генерирует Excel файл с бронированиями на текущий день
        
        Returns:
            Путь к сгенерированному файлу
        """
        try:
            # Создаем временный файл
            with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp_file:
                file_path = tmp_file.name
            
            # Создаем workbook и worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Бронирования"
            
            # Заголовки колонок
            headers = [
                "ID", "Время", "Гости", "Имя клиента", "Телефон",
                "Статус", "ID пользователя", "Создано", "Примечания"
            ]
            
            # Стили для заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Записываем заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Записываем данные
            for row, reservation in enumerate(reservations, 2):
                # Форматируем время
                reservation_time = reservation['reservation_time']
                if isinstance(reservation_time, str):
                    time_str = reservation_time
                else:
                    time_str = reservation_time.strftime("%H:%M")
                
                # Форматируем дату создания
                created_at = reservation['created_at']
                if isinstance(created_at, datetime):
                    created_str = created_at.strftime("%d.%m.%Y %H:%M")
                else:
                    created_str = str(created_at)
                
                # Статус с русским текстом
                status_map = {
                    'pending': '⏳ Ожидание',
                    'confirmed': '✅ Подтверждено', 
                    'cancelled': '❌ Отменено',
                    'completed': '🎉 Завершено'
                }
                status_text = status_map.get(reservation['status'], reservation['status'])
                
                data_row = [
                    reservation['id'],
                    time_str,
                    reservation['guests_count'],
                    reservation['customer_name'],
                    reservation['customer_phone'],
                    status_text,
                    reservation['user_id'],
                    created_str,
                    reservation.get('notes', '')
                ]
                
                for col, value in enumerate(data_row, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = thin_border
                    
                    # Центрируем числовые колонки
                    if col in [1, 3, 7]:  # ID, Гости, User ID
                        cell.alignment = Alignment(horizontal="center")
            
            # Настраиваем ширину колонок
            column_widths = [8, 10, 8, 20, 15, 15, 12, 16, 25]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # Добавляем фильтры
            ws.auto_filter.ref = f"A1:I{len(reservations) + 1}"
            
            # Добавляем сводную информацию
            summary_row = len(reservations) + 3
            ws.cell(row=summary_row, column=1, value="Сводка:").font = Font(bold=True)
            ws.cell(row=summary_row, column=2, value=f"Всего бронирований: {len(reservations)}")
            
            # Статистика по статусам
            status_counts = {}
            for reservation in reservations:
                status = reservation['status']
                status_counts[status] = status_counts.get(status, 0) + 1
            
            for i, (status, count) in enumerate(status_counts.items(), 1):
                status_text = status_map.get(status, status)
                ws.cell(row=summary_row + i, column=2, value=f"{status_text}: {count}")
            
            # Сохраняем файл
            wb.save(file_path)
            self.logger.info(f"✅ Generated reservations report with {len(reservations)} entries")
            
            return file_path
            
        except Exception as e:
            self.logger.error(f"❌ Failed to generate reservations report: {e}")
            raise
    
    def cleanup_file(self, file_path: str):
        """Очищает временный файл"""
        try:
            if os.path.exists(file_path):
                os.unlink(file_path)
                self.logger.debug(f"🧹 Cleared temporary file: {file_path}")
        except Exception as e:
            self.logger.warning(f"⚠️ Could not delete temporary file {file_path}: {e}")